from openpyxl import load_workbook, Workbook

filename = "evaluation_values.xlsx"

try:
    wb = load_workbook(filename)
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    ws.append(["CODE", "PRECISIONBOX", "RECALLBOX","F1BOX","PRECISIONSEGMENT","RECALLSEGMENT","F1SEGMENT", "MEANIOUBOX", "MEANIOUSEGMENT" "TIMEFLAT", "TIMEFLATNESTED" "TIMENESTED", "TYPE", "GT_BOX", "PRED_BOX", "PAIRS_BOX", "GT_SEGMENT", "PRED_SEGMENT", "PAIRS_SEGMENT", "DP2_INDEX", "VOCAB_GROUNDTRUTH","VOCAB_FRSTLVL", "VOCAB_SECONDLVL" ])

code = CODE
precisionbox = PRECISIONBOX
recallbox = RECALLBOX
f1box = F1BOX
precisionsegment = PRECISIONSEGMENT
recallsegment = RECALLSEGMENT
f1segment = F1SEGMENT
ioubox = MEANIOUBOX
iousegment = MEANIOUSEGMENT
timeflat = TIMEFLAT 
timeflatnested = TIMEFLATNESTED
timenested = TIMENESTED
type_ = TYPE 
gt_box = GT_BOX
pred_box = PRED_BOX
pairs_box = PAIRS_BOX
gt_segment = GT_SEGMENT
pred_segment = PRED_SEGMENT
pairs_segment = PAIRS_SEGMENT
dp2_index = DP2_INDEX
voc_groundtruth = VOCAB_GROUNDTRUTH
voc_firstlvl = VOCAB_FRSTLVL
voc_secondlvl = VOCAB_SECONDLVL


already_exists = False
for row in ws.iter_rows(min_row=2, values_only=True):
    code_cell = row[0]
    type_cell = row[12]
    if code_cell == code and type_cell == type_:
        already_exists = True
        break

if not already_exists:
    ws.append([code, precisionbox, recallbox, f1box,precisionsegment,recallsegment,f1segment, ioubox, iousegment, timeflat, timeflatnested, timenested, type_, gt_box, pred_box, pairs_box, gt_segment, pred_segment, pairs_segment, dp2_index, voc_groundtruth, voc_firstlvl, voc_secondlvl])

wb.save(filename)
